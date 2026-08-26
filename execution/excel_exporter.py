import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from typing import List, Dict, Any

def generate_excel_bytes(items: List[Dict[str, Any]]) -> io.BytesIO:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    
    headers = [
        "Item Type",
        "Manufacturer",
        "UPC",
        "Supplier SKUs",
        "SKU",
        "Name",
        "Short Name",
        "Price",
        "Cost",
        "Active",
        "Allow Cost Override",
        "Location Scope",
        "Location"
    ]
    ws.append(headers)
    
    # Header styling
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
    
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(1, col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        
    for item in items:
        supplier_sku = item.get("supplier_sku") or f"Hitfar:{item.get('hitfar_sku', '')}"
        price_val = item.get("price")
        cost_val = item.get("cost")
        
        row_data = [
            item.get("item_type", "Accessories - Cases"),
            item.get("manufacturer", "HyperGear"),
            item.get("upc", "-"),
            supplier_sku,
            item.get("sku", "-"),
            item.get("name", ""),
            item.get("short_name", "-"),
            price_val if price_val is not None else "",
            cost_val if cost_val is not None else "",
            item.get("active", 1),
            item.get("allow_cost_override", 1),
            item.get("location_scope", "global"),
            item.get("location", "")
        ]
        ws.append(row_data)
        
    # Auto-adjust column widths
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = openpyxl.utils.get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 3, 12)
        
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output
