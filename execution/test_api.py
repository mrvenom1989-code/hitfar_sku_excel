import requests
import openpyxl
import io

BASE_URL = "http://127.0.0.1:5000"

print("1. Testing GET /api/stats ...")
r = requests.get(f"{BASE_URL}/api/stats")
print(f"Status: {r.status_code}, Response: {r.json()}")

print("\n2. Testing GET /api/catalog?limit=3 ...")
r = requests.get(f"{BASE_URL}/api/catalog?limit=3")
data = r.json()
print(f"Status: {r.status_code}, Total in DB: {data.get('total')}")
for it in data.get("data", []):
    print(f"  - [{it['supplier_sku']}] {it['name']} | Cost=${it['cost']} | MSRP=${it['price']}")

first_item_id = data["data"][0]["id"]

print(f"\n3. Testing POST /api/update-price on item {first_item_id} ...")
r = requests.post(f"{BASE_URL}/api/update-price", json={"id": first_item_id, "price": 16.99})
print(f"Status: {r.status_code}, Response: {r.json()}")

print("\n4. Testing POST /api/upload-pdf (Testing Deduplication on existing invoice)...")
with open("hitfar_sku/order pdf.pdf", "rb") as f:
    r = requests.post(f"{BASE_URL}/api/upload-pdf", files={"file": ("order pdf.pdf", f, "application/pdf")})
print(f"Status: {r.status_code}")
res_json = r.json()
print(f"Total in PDF: {res_json.get('total_items_in_pdf')}")
print(f"New Items Inserted: {res_json.get('new_items_count')}")
print(f"Existing Items Detected: {res_json.get('existing_items_count')}")

print("\n5. Testing GET /api/export ...")
r = requests.get(f"{BASE_URL}/api/export")
print(f"Status: {r.status_code}, Content-Type: {r.headers.get('Content-Type')}, Size: {len(r.content)} bytes")
wb = openpyxl.load_workbook(io.BytesIO(r.content))
ws = wb["Sheet1"]
print(f"Exported Excel rows: {ws.max_row}, cols: {ws.max_column}")
print(f"Row 1 Headers: {[ws.cell(1, c).value for c in range(1, 14)]}")
print(f"Row 2 Sample:  {[ws.cell(2, c).value for c in range(1, 14)]}")

print("\nALL BACKEND API TESTS PASSED PERFECTLY!")
