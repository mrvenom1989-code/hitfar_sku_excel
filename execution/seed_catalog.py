import os
import sys
import openpyxl
import datetime

sys.path.append(os.path.abspath(os.path.join(os.path.dirname(__file__), "..")))
from execution.db_service import insert_catalog_items, get_catalog_stats

def seed():
    excel_path = os.path.join(os.path.dirname(os.path.dirname(__file__)), "order sku.xlsx")
    print(f"Reading seed data from {excel_path}...")
    
    wb = openpyxl.load_workbook(excel_path)
    ws = wb["Sheet1"]
    
    items = []
    today_str = datetime.date.today().isoformat()
    
    for row_idx in range(2, ws.max_row + 1):
        item_type = ws.cell(row_idx, 1).value
        manufacturer = ws.cell(row_idx, 2).value
        upc = ws.cell(row_idx, 3).value
        supplier_sku = ws.cell(row_idx, 4).value
        sku = ws.cell(row_idx, 5).value
        name = ws.cell(row_idx, 6).value
        short_name = ws.cell(row_idx, 7).value
        price = ws.cell(row_idx, 8).value
        cost = ws.cell(row_idx, 9).value
        active = ws.cell(row_idx, 10).value or 1
        allow_cost_override = ws.cell(row_idx, 11).value or 1
        location_scope = ws.cell(row_idx, 12).value or "global"
        location = ws.cell(row_idx, 13).value or ""
        
        if not supplier_sku or supplier_sku == "-":
            continue
            
        hitfar_sku = str(supplier_sku).replace("Hitfar:", "").strip()
        
        # Convert numeric prices
        try:
            price_val = float(price) if price is not None and str(price).strip() != "" else None
        except (ValueError, TypeError):
            price_val = None
            
        try:
            cost_val = float(cost) if cost is not None and str(cost).strip() != "" else None
        except (ValueError, TypeError):
            cost_val = None
            
        items.append({
            "item_type": item_type or "Accessories - Cases",
            "manufacturer": manufacturer or "HyperGear",
            "upc": upc or "-",
            "supplier_sku": supplier_sku,
            "hitfar_sku": hitfar_sku,
            "mpn": "",
            "sku": sku or "-",
            "name": name or "",
            "short_name": short_name or "-",
            "price": price_val,
            "cost": cost_val,
            "active": int(active),
            "allow_cost_override": int(allow_cost_override),
            "location_scope": location_scope,
            "location": location,
            "created_date": today_str,
            "last_order_po": "PO00001BP",
            "last_order_date": "2026-08-23"
        })
        
    print(f"Extracted {len(items)} items from spreadsheet. Inserting into database...")
    count = insert_catalog_items(items)
    print(f"Successfully inserted/updated {count} items!")
    
    stats = get_catalog_stats()
    print("Database Stats:", stats)

if __name__ == "__main__":
    seed()
